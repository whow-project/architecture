import os
import sys
import csv
import pandas as pd
from openpyxl import load_workbook, Workbook

def convert(file_in, folder_out):
    wb = load_workbook(file_in)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        
        file_out = os.path.join(folder_out, sheetname + '.csv')
        with open(file_out, 'wb') as f:   # open('a_file.csv', 'w', newline="") for python 3
            c = csv.writer(f, delimiter=';')
            for r in sheet.iter_rows():
                print([cell.value for cell in r])
                c.writerow([cell.value for cell in r])
                
            print("Exported %s"%file_out)
            
def convert_pd(file_in, folder_out):
    print ('Reading input file ...')
    df = pd.read_excel(file_in, None)
    for sheetname in df.keys():
        print ('Creating', sheetname, '...')
        file_out = os.path.join(folder_out, sheetname + '.csv')
        df[sheetname].to_csv(file_out, sep=';', index=None)
            
if __name__ == '__main__':
    file_in = sys.argv[1]
    folder_out = os.path.join('data', 'bathw', 'v2', 'dirtydata')
    
    if not os.path.exists(folder_out):
        os.makedirs(folder_out)
        print("Created folder %s"%folder_out) 
    
    convert_pd(file_in, folder_out)
