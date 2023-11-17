#Usage: python3 ostreopsis_excel_to_csv.py <input_xls_file.xlsx>

import os
import sys
import csv
import pandas as pd
from openpyxl import load_workbook

def convert(file_in, folder_out):
    wb = load_workbook(file_in)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        
        file_out = os.path.join(folder_out, sheetname + '.csv')
        with open(file_out, 'wb') as f:
            c = csv.writer(f, delimiter=';')
            for r in sheet.iter_rows():
                print([cell.value for cell in r])
                c.writerow([cell.value for cell in r])
                
            print("Exported %s"%file_out)
            
def convert_pd(year, file_in, folder_out):
    print ('Reading input file ...')
    df = pd.read_excel(file_in, None);
    for sheetname in df.keys():
        print ('Creating', sheetname, '...')
        file_out = os.path.join(folder_out, sheetname + '_'+year+'.csv')
        df[sheetname].to_csv(file_out, sep=';', index=None)
            
if __name__ == '__main__':
    file_in = sys.argv[1]
    YY = file_in.split('.')[0].split('_')[-1]
    folder_out = os.path.join('data', 'ostreopsis', 'csv')
    
    if not os.path.exists(folder_out):
        os.makedirs(folder_out)
        print("Created folder %s"%folder_out) 
    
    convert_pd(YY, file_in, folder_out)
